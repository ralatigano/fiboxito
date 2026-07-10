"""Registro de comprobantes en Excel sobre el NAS.

La fuente de verdad es la tabla `comprobantes_procesados` (SQLite). El .xlsx del
NAS es un export REGENERABLE: cada vez que entra un comprobante se reconstruye
el archivo completo desde la base y se sube. Así nunca se corrompe por
escrituras parciales y siempre es consistente con la base.

Manejo de "archivo abierto por alguien":
- Cuando alguien abre el .xlsx en Excel (por SMB), Excel crea un archivo de
  bloqueo `~$<nombre>.xlsx` en la misma carpeta. Antes de subir chequeamos si
  existe: si está, NO pisamos el archivo (para no chocar con quien lo edita) y
  dejamos el export "pendiente". Como la base ya tiene el dato, el próximo ciclo
  reintenta y el Excel se pone al día sin perder nada.
- Cuando no está bloqueado, subimos de forma ATÓMICA: escribimos a un temporal y
  lo renombramos encima del definitivo, así nadie llega a leer un archivo a
  medio escribir.
"""

import io
import os
import posixpath
import logging
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from db import get_conn
from routers.nas import service as nas
from routers.nas.sftp_client import NASError

log = logging.getLogger("fiboxito")

# Ruta del Excel dentro de la base del NAS (relativa a NAS_BASE_PATH).
# Fibox es la única carpeta con permiso de escritura de FiboxitoBot.
XLSX_PATH = os.getenv("NAS_COMPROBANTES_XLSX", "Fibox/Comprobantes/Comprobantes.xlsx")

# Diferencia horaria para mostrar "Registrado" en hora local (Argentina UTC-3).
TZ_OFFSET_HORAS = int(os.getenv("TZ_OFFSET_HORAS", "-3"))

# (encabezado, ancho de columna)
COLUMNAS = [
    ("Comprobante",   16),
    ("Fecha",         12),
    ("Mes",           10),
    ("Resultado",     11),
    ("Motivo",        45),
    ("Envió",         26),
    ("Pagó",          26),
    ("Monto",         14),
    ("N° operación",  30),
    ("Link",          10),
    ("Registrado",    18),
]

# Reintento diferido: si el archivo estaba abierto, queda pendiente.
_export_pendiente = False


def hay_export_pendiente() -> bool:
    return _export_pendiente


def _mes_de(fecha_str: str | None) -> str:
    """De 'dd/mm/aaaa' o 'dd-mm-aaaa' devuelve 'aaaa-mm' para filtrar por mes."""
    if not fecha_str:
        return ""
    txt = fecha_str.strip().replace("-", "/")
    partes = txt.split("/")
    if len(partes) == 3:
        try:
            dia, mes, anio = (int(p) for p in partes)
            return f"{anio:04d}-{mes:02d}"
        except ValueError:
            return ""
    return ""


def _registrado_local(iso_utc: str | None) -> str:
    if not iso_utc:
        return ""
    try:
        dt = datetime.fromisoformat(iso_utc) + timedelta(hours=TZ_OFFSET_HORAS)
        return dt.strftime("%d/%m/%Y %H:%M")
    except ValueError:
        return iso_utc


def _leer_filas() -> list:
    with get_conn() as conn:
        return conn.execute("""
            SELECT protocolo, fecha, resultado, motivo, contacto_wa, emisor,
                   monto, numero_op, url_imagen, procesado_at
            FROM comprobantes_procesados
            ORDER BY procesado_at ASC
        """).fetchall()


def _construir_xlsx(filas) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comprobantes"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    # Encabezados
    for c, (titulo, ancho) in enumerate(COLUMNAS, start=1):
        cell = ws.cell(row=1, column=c, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = ancho

    # Filas de datos
    for r, fila in enumerate(filas, start=2):
        ws.cell(row=r, column=1, value=fila["protocolo"])
        ws.cell(row=r, column=2, value=fila["fecha"])
        ws.cell(row=r, column=3, value=_mes_de(fila["fecha"]))
        ws.cell(row=r, column=4, value=fila["resultado"])
        ws.cell(row=r, column=5, value=fila["motivo"])
        ws.cell(row=r, column=6, value=fila["contacto_wa"])
        ws.cell(row=r, column=7, value=fila["emisor"])
        monto_cell = ws.cell(row=r, column=8, value=fila["monto"])
        if fila["monto"] is not None:
            monto_cell.number_format = '"$"#,##0.00'
        ws.cell(row=r, column=9, value=fila["numero_op"])
        # Link al comprobante como hipervínculo
        url = fila["url_imagen"]
        link_cell = ws.cell(row=r, column=10, value="Ver" if url else None)
        if url:
            link_cell.hyperlink = url
            link_cell.font = Font(color="0563C1", underline="single")
        ws.cell(row=r, column=11, value=_registrado_local(fila["procesado_at"]))

    # Filtros + panel fijo en el encabezado
    ultima_col = get_column_letter(len(COLUMNAS))
    ws.auto_filter.ref = f"A1:{ultima_col}{max(ws.max_row, 1)}"
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _lockfile(path: str) -> str:
    """Ruta del archivo de bloqueo de Excel (~$nombre.xlsx) en la misma carpeta."""
    carpeta = posixpath.dirname(path)
    nombre = posixpath.basename(path)
    return posixpath.join(carpeta, f"~${nombre}")


def exportar() -> dict:
    """Regenera el Excel desde la base y lo sube al NAS.

    Devuelve {"estado": "ok"|"diferido"|"error", "detalle": str}.
    """
    global _export_pendiente
    try:
        # ¿Alguien lo tiene abierto? No pisamos su edición.
        if nas.existe(_lockfile(XLSX_PATH)):
            _export_pendiente = True
            log.warning("[EXPORT] El Excel está abierto por alguien — export diferido.")
            return {"estado": "diferido",
                    "detalle": "El archivo está abierto; se actualizará automáticamente al cerrarlo."}

        contenido = _construir_xlsx(_leer_filas())

        # Escritura atómica: subir a temporal y renombrar encima.
        tmp = f"{XLSX_PATH}.tmp"
        nas.escribir(tmp, contenido)
        nas.renombrar(tmp, XLSX_PATH)

        _export_pendiente = False
        log.info(f"[EXPORT] Excel actualizado en NAS: {XLSX_PATH}")
        return {"estado": "ok", "detalle": XLSX_PATH}

    except NASError as e:
        _export_pendiente = True
        log.error(f"[EXPORT] Error de NAS al exportar: {e}")
        return {"estado": "error", "detalle": str(e)}
    except Exception as e:  # noqa: BLE001
        _export_pendiente = True
        log.error(f"[EXPORT] Error inesperado al exportar: {e}")
        return {"estado": "error", "detalle": f"{type(e).__name__}: {e}"}
